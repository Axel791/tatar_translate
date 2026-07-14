#!/usr/bin/env python3
"""Конвейер перевода достижений (RU → TT). Работа в 4 машины, каждая со своей частью.

Команды:
  status [N]                    — общий прогресс (или прогресс части N)
  slice N P [NB] [SIZE]         — нарезать порцию P для машины N: NB батчей (деф. 24) по SIZE (деф. 25)
                                  непереведённых текстов из work/parts/part_N.jsonl -> work/queue/mNpP_bXX.json
  merge                         — влить work/batches/*.jsonl в чекпоинт с проверками -> отчёт
  apply                         — записать переводы в лист «Перевод достижений» файла Пакет200.xlsx (координатор)
  apply-part N                  — заполнить колонку переводов в work/parts/part_N.xlsx (локальный контроль)
"""
import json, sys, os, re, glob, hashlib

WORK = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(WORK)
UNIQUE = os.path.join(WORK, 'dost_unique.jsonl')
CHECKPOINT = os.path.join(WORK, 'dost_translations.jsonl')
PARTS = os.path.join(WORK, 'parts')
QUEUE = os.path.join(WORK, 'queue')
BATCHES = os.path.join(WORK, 'batches')
XLSX = os.path.join(ROOT, 'Пакет200.xlsx')

# слова, которые обязаны быть переведены (русизмы-маркеры халтуры).
FORBIDDEN = ['награжден', 'награждена', 'почетная грамота', 'почетной грамотой',
             'заслуженный', 'заслуженная', 'благодарностью', 'благодарственное',
             'работает', 'работник', 'сотрудник', 'зарекомендовал', 'зарекомендовала',
             'опыт работы', 'принимал участие', 'принимала участие', 'является',
             'руководством', 'звание', 'учитель', 'воспитатель', 'директор школы']

def load_unique():
    return [json.loads(l) for l in open(UNIQUE, encoding='utf-8')]

def load_part(n):
    return [json.loads(l) for l in open(os.path.join(PARTS, f'part_{n}.jsonl'), encoding='utf-8')]

def load_checkpoint():
    done = {}
    if os.path.exists(CHECKPOINT):
        for l in open(CHECKPOINT, encoding='utf-8'):
            l = l.strip()
            if not l:
                continue
            o = json.loads(l)
            done[o['k']] = o['tt']
    return done

def cmd_status(part=None):
    done = load_checkpoint()
    if part:
        items = load_part(part)
        d = sum(1 for it in items if it['k'] in done)
        print(f"Часть {part}: переведено {d} из {len(items)} ({100*d/len(items):.1f}%)")
        return
    items = load_unique()
    rows_total = rows_done = 0
    for it in items:
        rows_total += it['freq']
        if it['k'] in done:
            rows_done += it['freq']
    print(f"Уникальных текстов: {len(items)}, переведено: {len(done)} ({100*len(done)/len(items):.1f}%)")
    print(f"Строк таблицы покрыто: {rows_done} из {rows_total} ({100*rows_done/rows_total:.1f}%)")
    for n in (1, 2, 3, 4):
        if os.path.exists(os.path.join(PARTS, f'part_{n}.jsonl')):
            p = load_part(n)
            d = sum(1 for it in p if it['k'] in done)
            print(f"  часть {n}: {d}/{len(p)} ({100*d/len(p):.1f}%)")

def cmd_slice(machine, portion, nbatches=24, size=25):
    os.makedirs(QUEUE, exist_ok=True)
    done = load_checkpoint()
    todo = [it for it in load_part(machine) if it['k'] not in done]
    made = 0
    for i in range(nbatches):
        chunk = todo[i*size:(i+1)*size]
        if not chunk:
            break
        name = f'm{machine}p{portion}_b{i+1:02d}'
        with open(os.path.join(QUEUE, name + '.json'), 'w', encoding='utf-8') as f:
            json.dump([{'k': c['k'], 'ru': c['ru']} for c in chunk], f, ensure_ascii=False, indent=0)
        made += 1
    left = max(0, len(todo) - nbatches*size)
    print(f'Машина {machine}, порция {portion}: батчей {made} по {size} '
          f'(тексты: {min(len(todo), nbatches*size)}); в части останется: {left}')
    for i in range(made):
        print(f'  work/queue/m{machine}p{portion}_b{i+1:02d}.json')

DIGITS = re.compile(r'\d+')

def check_pair(ru, tt):
    """Возвращает список проблем для пары перевода."""
    probs = []
    if not tt or not tt.strip():
        probs.append('пустой перевод')
        return probs
    if len(ru.strip()) <= 3:
        return probs
    ru_d = DIGITS.findall(ru)
    tt_d = DIGITS.findall(tt)
    missing = [d for d in set(ru_d) if ru_d.count(d) > tt_d.count(d)]
    if len(missing) > 1:
        probs.append(f'потеряны числа: {sorted(missing)[:6]}')
    if len(tt) < 0.4 * len(ru):
        probs.append(f'подозрительно короткий ({len(tt)} vs {len(ru)})')
    if tt.count('«') != tt.count('»'):
        probs.append('кавычки « » не сбалансированы')
    low = re.sub(r'«[^»]*»', '«»', tt).lower()
    hits = [w for w in FORBIDDEN if w in low]
    if hits:
        probs.append(f'непереведённые русизмы: {hits[:4]}')
    return probs

def cmd_merge():
    items = {it['k']: it for it in load_unique()}
    done = load_checkpoint()
    new, bad, dup, unknown = {}, [], 0, 0
    for path in sorted(glob.glob(os.path.join(BATCHES, '*.jsonl'))):
        for ln, l in enumerate(open(path, encoding='utf-8'), 1):
            l = l.strip()
            if not l:
                continue
            try:
                o = json.loads(l)
                k, tt = o['k'], str(o['tt']).strip()
            except Exception as e:
                bad.append((path, ln, f'битый JSON: {e}'))
                continue
            if k not in items:
                unknown += 1
                bad.append((path, ln, f'неизвестный ключ {k}'))
                continue
            if k in done or k in new:
                dup += 1
                continue
            probs = check_pair(items[k]['ru'], tt)
            if probs:
                bad.append((path, ln, f'k={k}: ' + '; '.join(probs)))
                continue
            new[k] = tt
    with open(CHECKPOINT, 'a', encoding='utf-8') as f:
        for k, tt in new.items():
            f.write(json.dumps({'k': k, 'tt': tt}, ensure_ascii=False) + '\n')
    print(f'Влито новых переводов: {len(new)}; дублей пропущено: {dup}; проблемных: {len(bad)}')
    for p, ln, msg in bad[:40]:
        print(f'  ПРОБЛЕМА {os.path.basename(p)}:{ln} {msg}')
    if len(bad) > 40:
        print(f'  ... и ещё {len(bad)-40}')
    if bad:
        keys = set()
        for p, ln, msg in bad:
            m = re.search(r'k=([0-9a-f]{12})', msg)
            if m:
                keys.add(m.group(1))
        with open(os.path.join(WORK, 'retry_keys.txt'), 'w') as f:
            f.write('\n'.join(sorted(keys)))
        print(f'Ключи на повтор: {len(keys)} -> work/retry_keys.txt '
              f'(повторный slice подхватит их автоматически)')

def cmd_apply():
    import openpyxl
    done = load_checkpoint()
    wb = openpyxl.load_workbook(XLSX)
    name = 'Перевод достижений'
    src = wb['Лауреаты']
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
        ws['A1'], ws['B1'], ws['C1'] = 'ID лауреата', 'Достижения', 'Достижения (тат.)'
    filled = empty_src = 0
    for i, r in enumerate(src.iter_rows(min_row=2, values_only=True), start=2):
        rid, dol, dost = (r + (None,)*3)[:3]
        ws.cell(row=i, column=1, value=rid)
        ws.cell(row=i, column=2, value=dost)
        t = str(dost).strip() if dost and str(dost).strip() else None
        if not t:
            empty_src += 1
            continue
        k = hashlib.sha1(t.encode('utf-8')).hexdigest()[:12]
        tt = done.get(k)
        if tt:
            ws.cell(row=i, column=3, value=tt)
            filled += 1
    wb.save(XLSX)
    print(f'Заполнено переводов: {filled}; пустых исходников: {empty_src}; строк всего: {src.max_row-1}')

def cmd_apply_part(n):
    import openpyxl
    done = load_checkpoint()
    path = os.path.join(PARTS, f'part_{n}.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    filled = 0
    for row in ws.iter_rows(min_row=2):
        k = row[1].value
        if k and k in done:
            row[3].value = done[k]
            filled += 1
    wb.save(path)
    print(f'part_{n}.xlsx: заполнено {filled} переводов')

if __name__ == '__main__':
    cmd = sys.argv[1] if len(sys.argv) > 1 else 'status'
    if cmd == 'status':
        cmd_status(int(sys.argv[2]) if len(sys.argv) > 2 else None)
    elif cmd == 'slice':
        cmd_slice(int(sys.argv[2]), int(sys.argv[3]),
                  int(sys.argv[4]) if len(sys.argv) > 4 else 24,
                  int(sys.argv[5]) if len(sys.argv) > 5 else 25)
    elif cmd == 'merge':
        cmd_merge()
    elif cmd == 'apply':
        cmd_apply()
    elif cmd == 'apply-part':
        cmd_apply_part(int(sys.argv[2]))
    else:
        print(__doc__)
